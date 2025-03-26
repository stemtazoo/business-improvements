import os
import re

import flet as ft
from openpyxl import load_workbook, Workbook
import yaml

# デフォルトの設定を定義
DEFAULT_CONFIG = {
    "sheet_filename_format": "{sheet_name}.xlsx"
}

def load_config():
    """
    設定ファイルを読み込む。ファイルが存在しない場合はデフォルト設定を返す。
    """
    try:
        # 設定ファイルのパスを調整
        config_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 
            "..", 
            "config", 
            "app1_config.yaml"
        )
        
        # ファイルが存在する場合のみ読み込む
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f)
                return config
        else:
            print(f"設定ファイルが見つかりません: {config_path}")
            return DEFAULT_CONFIG
    except Exception as e:
        print(f"設定ファイルの読み込み中にエラーが発生しました: {e}")
        return DEFAULT_CONFIG

def sanitize_filename(name):
    return re.sub(r'[\\\\/:*?"<>|]', '_', name)

def split_excel_sheets(input_file: str, output_dir: str = None):
    """
    指定した Excel ファイルを開き、各シートを別ファイルとして出力する。
    選択したファイル名に基づいたフォルダを作成し、そこにファイルを出力する。

    :param input_file: 元となる Excel ファイルパス
    :param output_dir: 出力先ディレクトリ（None の場合は input_file と同じディレクトリ）
    """
    
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"指定されたファイルが見つかりません: {input_file}")

    wb = load_workbook(input_file)
    base_dir = os.path.dirname(input_file)

    if output_dir is None or output_dir.strip() == "":
        output_dir = base_dir

    # 入力ファイルの名前から出力フォルダ名を生成
    input_filename = os.path.splitext(os.path.basename(input_file))[0]
    output_folder_name = f"{input_filename}_分割したファイル"
    output_folder_path = os.path.join(output_dir, output_folder_name)

    # 出力フォルダが存在しない場合は作成
    os.makedirs(output_folder_path, exist_ok=True)

    # 設定ファイルを読み込み
    config = load_config()
    sheet_filename_format = config.get("sheet_filename_format", "{sheet_name}.xlsx")

    for sheet_name in wb.sheetnames:
        print(f"=== 分割処理中: {sheet_name} ===")
        sheet = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        # シート内容をコピー
        for row in sheet.iter_rows(values_only=True):
            new_ws.append(row)

        # 出力ファイル名を作成
        safe_name = sanitize_filename(sheet_name)
        output_file_name = sheet_filename_format.format(sheet_name=safe_name)
        output_file_path = os.path.join(output_folder_path, output_file_name)

        try:
            new_wb.save(output_file_path)
            print(f"✅ 保存成功: {output_file_path}")
        except Exception as e:
            print(f"❌ 保存失敗: {output_file_path}")
            print(f"エラー内容: {e}")


def main(page: ft.Page):
    # ウィンドウの設定
    page.title = "Excel シート分割ツール"
    
    page.window.width = 500  
    page.window.height = 550  
    page.window.min_width = 300
    page.window.min_height = 450
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.update()

    # ファイルパス用のテキストフィールド
    file_path_input = ft.TextField(
        label="選択されたファイルパス", 
        width=450, 
        read_only=True,
        hint_text="Excel ファイルを選択してください"
    )

    # 出力フォルダパス用のテキストフィールド
    output_folder_input = ft.TextField(
        label="出力先フォルダパス", 
        width=450, 
        read_only=True,
        hint_text="（省略可）出力先フォルダを選択"
    )

    # 結果表示用のテキスト
    result_text = ft.Text(color=ft.colors.GREEN)

    # ファイル選択ダイアログ
    def pick_file(e):
        file_picker.pick_files(
            dialog_title="Excelファイルを選択",
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx", "xlsm", "xltx", "xltm"]
        )

    # フォルダ選択ダイアログ
    def pick_folder(e):
        folder_picker.get_directory_path(dialog_title="出力先フォルダを選択")

    # ファイル選択結果のハンドリング
    def on_file_selected(e: ft.FilePickerResultEvent):
        if e.files:
            # ファイルパスを設定
            file_path = e.files[0].path
            file_path_input.value = file_path
            
            # 出力先フォルダをデフォルトでファイルと同じディレクトリに設定
            default_output_folder = os.path.dirname(file_path)
            output_folder_input.value = default_output_folder
            
            page.update()

    # フォルダ選択結果のハンドリング
    def on_folder_selected(e: ft.FilePickerResultEvent):
        if e.path:
            output_folder_input.value = e.path
            page.update()

    # シート分割処理
    def split_sheets(e):
        input_file = file_path_input.value
        output_folder = output_folder_input.value or None

        if not input_file:
            result_text.value = "Excel ファイルを選択してください。"
            result_text.color = ft.colors.RED
            page.update()
            return

        try:
            split_excel_sheets(input_file, output_folder)
            result_text.value = "処理が完了しました。"
            result_text.color = ft.colors.GREEN
        except Exception as ex:
            result_text.value = f"エラーが発生しました: {ex}"
            result_text.color = ft.colors.RED
        
        page.update()

    # ファイルピッカーとフォルダピッカーの設定
    file_picker = ft.FilePicker(on_result=on_file_selected)
    folder_picker = ft.FilePicker(on_result=on_folder_selected)
    page.overlay.extend([file_picker, folder_picker])

    # UIレイアウト
    page.add(
        ft.Column([
            ft.Text("Excel シート分割ツール", size=20, weight=ft.FontWeight.BOLD),
            
            ft.Text("1. Excel ファイルを選択"),
            ft.ElevatedButton(
                "ファイルを選択", 
                icon=ft.icons.UPLOAD_FILE, 
                on_click=pick_file
            ),
            file_path_input,

            ft.Text("2. 出力先フォルダを選択（省略可）"),
            ft.ElevatedButton(
                "フォルダを選択", 
                icon=ft.icons.FOLDER_OPEN, 
                on_click=pick_folder
            ),
            output_folder_input,

            ft.ElevatedButton(
                "シートを分割して保存", 
                icon=ft.icons.SAVE, 
                on_click=split_sheets
            ),
            result_text
        ], 
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        width=500,
        spacing=10,
        )
    )

    # ウィンドウを画面中央に配置する新しい方法
    page.window_top = None
    page.window_left = None

if __name__ == "__main__":
    ft.app(target=main)