import os
import pytest
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook
from src.main import split_excel_sheets

@pytest.fixture
def sample_excel_file(tmp_path):
    """テスト用のサンプルExcelファイルを作成して返す"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B", "C"])
    ws1.append([1, 2, 3])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y", "Z"])
    ws2.append([10, 20, 30])

    file_path = tmp_path / "test_input.xlsx"
    wb.save(file_path)
    return file_path

def test_split_excel_sheets_creates_files(tmp_path, sample_excel_file):
    output_dir = tmp_path / "output"
    output_dir.mkdir()

    split_excel_sheets(str(sample_excel_file), str(output_dir))
    print("=== 出力されたファイル一覧 ===")
    for f in output_dir.glob("*.xlsx"):
        print(f.name)


    sheet1_file = output_dir / "Sheet1.xlsx"
    sheet2_file = output_dir / "Sheet2.xlsx"

    assert sheet1_file.exists()
    assert sheet2_file.exists()

    # 内容確認
    wb1 = load_workbook(sheet1_file)
    ws1 = wb1.active
    assert ws1.title == "Sheet1"
    assert ws1.cell(row=2, column=1).value == 1

    wb2 = load_workbook(sheet2_file)
    ws2 = wb2.active
    assert ws2.title == "Sheet2"
    assert ws2.cell(row=2, column=3).value == 30
